""" Just a testing doc used to learn about the things after to use this in OOP.PY"""

import openpyxl
import pandas as pd
# Caminho para o arquivo Excel existente
CAMINHO_ARQUIVO_EXCEL = 'data_base.xlsx'

# Nome da planilha onde você deseja adicionar a linha
NOME_PLANILHA = 'Administradores'  # Altere para o nome da planilha desejada

# Carregar o arquivo Excel
arquivo_excel = openpyxl.load_workbook(CAMINHO_ARQUIVO_EXCEL)

# Selecionar a planilha desejada
planilha = arquivo_excel[NOME_PLANILHA]

# Encontrar a última linha preenchida na planilha
ultima_linha = planilha.max_row

# Adicionar uma nova linha após a última linha preenchida
nova_linha = ultima_linha + 1

# Adicionar os dados à nova linha
# Suponha que você tenha uma lista chamada 'dados' com os valores para cada coluna
dados = pd.DataFrame({'A': [1, 2, 3],
                    'B': [4, 5, 6],
                    'C': [7, 8, 9]})

print(pd.DataFrame.info(dados))

print(dados.loc[1])
print(dados.head())

#for coluna, valor in enumerate(dados, start=1):
#    planilha.cell(row=nova_linha, column=coluna, value=valor)
#
# Salvar as alterações no arquivo Excel
#arquivo_excel.save(caminho_arquivo_excel)
#
# Fechar o arquivo Excel
#arquivo_excel.close()
